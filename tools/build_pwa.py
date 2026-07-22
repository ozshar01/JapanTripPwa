#!/usr/bin/env python3
import json, sys, os
from datetime import date, datetime, time
from openpyxl import load_workbook

ROOT=os.path.abspath(os.path.join(os.path.dirname(__file__),'..'))
DEFAULT_XLSX=os.path.join(ROOT,'source','trip-template.xlsx')
OUT=os.path.join(ROOT,'data','trip-data.js')

def val(v):
    if isinstance(v,(date,datetime)): return v.strftime('%Y-%m-%d')
    if isinstance(v,time): return v.strftime('%H:%M')
    return '' if v is None else v

def rows(ws):
    hdr=[str(c.value).strip() if c.value is not None else '' for c in ws[1]]
    out=[]
    for row in ws.iter_rows(min_row=2,values_only=True):
        if not any(x is not None and str(x).strip() for x in row): continue
        out.append({hdr[i]:val(row[i]) for i in range(min(len(hdr),len(row)))})
    return out

def parse_list(s):
    if not s: return []
    return [x.strip() for x in str(s).replace('\r','').split('\n') if x.strip()]

def truth(v): return str(v).strip().lower() in ('כן','true','1','yes','y')

def main():
    xlsx=sys.argv[1] if len(sys.argv)>1 else DEFAULT_XLSX
    wb=load_workbook(xlsx,data_only=True)
    required=['הגדרות','ימים','לוח זמנים','הזמנות','מסעדות','מלונות','משפטים']
    missing=[s for s in required if s not in wb.sheetnames]
    if missing: raise SystemExit('Missing sheets: '+', '.join(missing))
    settings={r['שדה']:r['ערך'] for r in rows(wb['הגדרות'])}
    days=[]
    for r in rows(wb['ימים']):
        days.append({'id':r['מזהה'],'day':int(r['יום']),'date':r['תאריך'],'weekday':r['יום בשבוע'],'city':r['עיר'],'title':r['כותרת'],'subtitle':r['כותרת משנה'],'hotelId':r['מזהה מלון'],'load':r['עומס'],'holiday':truth(r['חג']),'important':parse_list(r['חשוב לדעת']),'take':parse_list(r['מה לקחת'])})
    schedule=[]
    for r in rows(wb['לוח זמנים']):
        schedule.append({'id':r['מזהה'],'dayId':r['מזהה יום'],'order':int(r['סדר']),'start':r['שעת התחלה'],'end':r['שעת סיום'],'title':r['כותרת'],'description':r['תיאור'],'transport':r['תחבורה'],'place':r['מקום'],'mapsQuery':r['חיפוש במפות'],'bookingId':r['מזהה הזמנה'],'internet':truth(r['דורש אינטרנט'])})
    bookings=[]
    for r in rows(wb['הזמנות']):
        bookings.append({'id':r['מזהה'],'dayId':r['מזהה יום'],'name':r['שם'],'status':r['סטטוס'],'priority':r['עדיפות'],'deadline':r['תאריך יעד'],'useDate':r['תאריך שימוש'],'time':r['שעה'],'arrival':r['הגעה מומלצת'],'cost':r['עלות משוערת'],'url':r['קישור'],'notes':r['הערות']})
    restaurants=[]
    for r in rows(wb['מסעדות']):
        restaurants.append({'id':r['מזהה'],'days':parse_list(r['ימים']),'area':r['אזור'],'name':r['שם'],'cuisine':r['מטבח'],'price':r['מחיר'],'porkStatus':r['התאמה ללא חזיר'],'recommended':r['מומלץ להזמין'],'mapsQuery':r['חיפוש במפות'],'website':r['אתר'],'notes':r['הערות']})
    hotels=[]
    for r in rows(wb['מלונות']):
        hotels.append({'id':r['מזהה'],'name':r['שם'],'city':r['עיר'],'address':r['כתובת'],'phone':r['טלפון'],'mapsQuery':r['חיפוש במפות'],'notes':r['הערות']})
    phrases=[{'category':r['קטגוריה'],'he':r['עברית'],'ja':r['יפנית'],'roman':r['תעתיק']} for r in rows(wb['משפטים'])]
    data={'version':datetime.now().strftime('%Y-%m-%d %H:%M'),'settings':settings,'days':days,'schedule':schedule,'bookings':bookings,'restaurants':restaurants,'hotels':hotels,'phrases':phrases}
    os.makedirs(os.path.dirname(OUT),exist_ok=True)
    with open(OUT,'w',encoding='utf-8') as f: f.write('window.TRIP_DATA = '+json.dumps(data,ensure_ascii=False,indent=2)+';\n')
    with open(os.path.join(ROOT,'data','trip-data.json'),'w',encoding='utf-8') as f: json.dump(data,f,ensure_ascii=False,indent=2)
    print(f'Built {OUT}: {len(days)} days, {len(schedule)} activities, {len(restaurants)} restaurants')
if __name__=='__main__': main()
